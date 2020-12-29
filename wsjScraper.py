import tkinter as tk
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl
import time
from datetime import date

def create_error_window(error_message,width,height):
    # create root window
    root = tk.Tk()
    # width and height of user scree
    screenW = root.winfo_screenwidth()
    screenH = root.winfo_screenheight()
    # x and y coords of center of screen
    x = (screenW/2) - (width/2)    
    y = (screenH/2) - (height/2)
    # set root starting position and size
    root.geometry('%dx%d+%d+%d' % (width, height, x, y))
    # not resizable
    root.resizable(0,0)
    # title
    root.title('Error')
    text = tk.Text(root)
    text.pack()
    # error message
    text.insert(tk.END,error_message)
    # launch window
    root.mainloop() 

def write_vals_to_xl(vals):
    file_path = r'C:/Users/gbonn/coding_Projects/stocksWebScraper/newPerformanceData.xlsx'
    # load excel spreadsheet
    try:
        wb = openpyxl.load_workbook(file_path)
    except:
        loading_error = "An error has occurred.\n\nThe spreadsheet cannot be found at path:\n" + file_path
        create_error_window(loading_error,400,100)
        return
    # create a backup of current spreadsheet
    try:
        backup_file_path = r'C:/Users/gbonn/coding_Projects/stocksWebScraper/backup/testWorksheetBackup.xlsx'
        wb.save(backup_file_path)
    except:
        backup_error = "An error has occurred.\n\nThe backup file cannot be saved to path:\n" + backup_file_path
        create_error_window(backup_error,400,100)
    # add new values to current spreadsheet
    ws = wb['Sheet1']
    currRow = 8
    currCell = ws.cell(currRow,2)
    currCell.value = date.today()
    curr_indx = 0
    for i in range(9,9+len(vals)+3):
        if(i != 14 and i != 22 and i != 25):
            currCell = ws.cell(i,2) 
            currCell.value = vals[curr_indx]
            curr_indx = curr_indx + 1
    # save updated copy of spreadsheet 
    try:
        wb.save(file_path)
    except:
        driver_error = "An error has occurred.\n\nThe spreadsheet cannot be saved.\n\nTry saving and closing the spreadsheet and\nrunning the program again."
        create_error_window(driver_error,400,100)

def searchWSJ(driver):
    # url base for searches
    url_base = 'https://www.wsj.com/search?query='
    # url extensions for what will be entered into the search bar 
    url_end = ['DJIA','SPX','COMP','RUT','RMCC','UKX','SXXP','NIK','SHCOMP','BVSP','891800','BUXX','GC00','CL.1']
    vals_found = []
    # go to this website
    for indx in url_end:
        driver.get(url_base+indx)
        # loading remains False while driver is loading
        loading = False
        # current amount of seconds waited for page to load
        safety_catch = 0
        while(loading == False):
            # try to get value if page has loaded 
            try:
                curr_val = driver.find_element_by_class_name('WSJTheme--last--3GifPA1e ').get_attribute("innerHTML")
                vals_found.append(round(float(curr_val),2))
                loading = True
            except:
                # if value isn't there wait one second and restart loop
                time.sleep(1)
                loading = False
                safety_catch = safety_catch + 1
                # if waited more than 9 seconds for load break loop
                if(safety_catch > 9):
                    vals_found.append('Not Found')
                    break
    
    return vals_found
    

def searchTreasury(driver):
    # url to treasury website
    url = 'https://www.treasury.gov/resource-center/data-chart-center/interest-rates/Pages/TextView.aspx?data=yield'
    driver.get(url)
    vals_found = []
    # find where the elements are 
    oddrows = driver.find_elements_by_class_name('oddrow')
    last_row = oddrows[-1]
    td_list = last_row.find_elements_by_tag_name('td')
    # indexes of interest correspinding to 3M, 2Y, 10Y, and 30Y
    td_indxs = [3,7,11,12]
    for indx in td_indxs:
        curr_td = td_list[indx]
        curr_val = curr_td.get_attribute('innerHTML')
        vals_found.append(float(curr_val))
    return vals_found

def checkExcelClosed():
    root = tk.Tk()
    # width and height of window
    width = 400
    height = 100
    # width and height of user screen
    screenW = root.winfo_screenwidth()
    screenH = root.winfo_screenheight()
    # x and y coords of center of screen
    x = (screenW/2) - (width/2)    
    y = (screenH/2) - (height/2)
    # set root starting position and size
    root.geometry('%dx%d+%d+%d' % (width, height, x, y))
    # add text and button
    root.title('Close Excel')
    tk.Label(root,text='Click OK when you have saved and closed your current excel file.\n').pack(side=tk.TOP)
    button = tk.Button(root,text='OK',command = root.destroy,height=1,width=10)
    button.pack(side=tk.TOP)
    # display window
    root.mainloop()


def main():

    # check to make sure excel sheet is saved and closed
    checkExcelClosed()

    # try to find the chrome driver and launch it
    try:
        driver_path = r'C:/ChromeDriver/chromedriver.exe'
        driver = webdriver.Chrome(executable_path=driver_path)
    except:
        driver_error = "An error has occurred.\n\nchromedriver.exe not found in file path given:\n" + driver_path
        create_error_window(driver_error,400,100)
        return
    
    # search the Wall Street Journal for values
    wsj_vals = searchWSJ(driver)
    # search the treasury for values
    treas_vals = searchTreasury(driver)
    # write values to excel 
    combined_vals = wsj_vals + treas_vals
    write_vals_to_xl(combined_vals)
    
# launch program 
main()